from __future__ import annotations

import html
import json
import re
from dataclasses import dataclass
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.models import DocumentInfo, PdfInsight
from app.core.settings import settings


_MAX_LINE_CHARS = 480


@dataclass
class DocClassification:
    is_application_form: bool
    confidence: float
    reason: str


def _clean_text(text: str) -> str:
    value = html.unescape(text)
    value = value.replace("\u2002", " ").replace("\u00a0", " ")
    value = re.sub(r"\s+", " ", value)
    value = value.strip()
    if len(value) > _MAX_LINE_CHARS:
        return f"{value[:_MAX_LINE_CHARS]}..."
    return value


def _docx_lines(path: Path, max_lines: int) -> list[str]:
    with ZipFile(path) as zf:
        xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    xml = re.sub(r"</w:(?:p|tr|tc)>", "\n", xml)
    xml = re.sub(r"<[^>]+>", "", xml)
    lines: list[str] = []
    for row in xml.splitlines():
        text = _clean_text(row)
        if not text:
            continue
        lines.append(text)
        if len(lines) >= max_lines:
            break
    return lines


def _xlsx_lines(path: Path, max_lines: int) -> list[str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    try:
        for ws in wb.worksheets:
            max_col = min(ws.max_column or 0, 24)
            if max_col <= 0:
                continue
            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row or 0,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ):
                values: list[str] = []
                for cell_value in row:
                    if cell_value is None:
                        continue
                    cleaned = _clean_text(str(cell_value))
                    if cleaned:
                        values.append(cleaned)
                    if len(values) >= 12:
                        break
                if not values:
                    continue
                line = _clean_text(f"[sheet:{ws.title}] " + " | ".join(values))
                if line:
                    lines.append(line)
                if len(lines) >= max_lines:
                    return lines
    finally:
        wb.close()
    return lines


def _pdf_lines(path: Path, max_lines: int) -> list[str]:
    reader = PdfReader(str(path))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        for row in text.splitlines():
            cleaned = _clean_text(row)
            if not cleaned:
                continue
            lines.append(cleaned)
            if len(lines) >= max_lines:
                return lines
    return lines


def extract_preview_lines(path: Path, kind: str, max_lines: int, max_chars: int) -> list[str]:
    try:
        if kind in {"docx", "docm"}:
            lines = _docx_lines(path, max_lines)
        elif kind == "xlsx":
            lines = _xlsx_lines(path, max_lines)
        elif kind == "pdf":
            lines = _pdf_lines(path, max_lines)
        else:
            lines = []
    except Exception:
        return []

    if max_chars <= 0:
        return lines
    trimmed: list[str] = []
    size = 0
    for line in lines:
        cost = len(line) + 1
        if trimmed and size + cost > max_chars:
            break
        trimmed.append(line)
        size += cost
    return trimmed


def _parse_json(text: str) -> dict:
    try:
        payload = json.loads(text)
        if isinstance(payload, dict):
            return payload
    except Exception:
        pass
    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        return {}
    try:
        payload = json.loads(match.group(0))
        if isinstance(payload, dict):
            return payload
    except Exception:
        return {}
    return {}


def _coerce_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes"}
    return False


def _coerce_confidence(value: object, fallback: float = 0.0) -> float:
    try:
        val = float(value)
    except Exception:
        val = fallback
    if val < 0:
        return 0.0
    if val > 1:
        return 1.0
    return round(val, 4)


class LLMDocClassifier:
    def __init__(self) -> None:
        self.enabled = bool(settings.openai_api_key)
        self._client = None
        if self.enabled:
            try:
                from openai import OpenAI

                self._client = OpenAI(
                    api_key=settings.openai_api_key,
                    timeout=settings.doc_classifier_timeout_sec,
                    max_retries=settings.doc_classifier_max_retries,
                )
            except Exception:
                self.enabled = False
                self._client = None

    def classify(
        self,
        *,
        doc: DocumentInfo,
        preview_lines: list[str],
        pdf_insight: PdfInsight | None,
    ) -> DocClassification:
        if not self.enabled or self._client is None:
            return DocClassification(False, 0.0, "LLM unavailable")

        pdf_summary = None
        if pdf_insight is not None:
            pdf_summary = {
                "title": pdf_insight.title,
                "sections": [s.get("title", "") for s in pdf_insight.sections[:40]],
                "criteria_lines": pdf_insight.criteria_lines[:40],
                "required_document_lines": pdf_insight.required_document_lines[:40],
            }

        payload = {
            "task": (
                "Decide whether this file is a tender application form that suppliers must fill "
                "(company info, declarations, pricing forms, compliance forms, etc.)."
            ),
            "document": {
                "name": doc.name,
                "kind": doc.kind,
                "relative_path": doc.relative_path,
            },
            "preview_lines": preview_lines,
            "pdf_summary": pdf_summary or {},
            "output_schema": {
                "is_application_form": "boolean",
                "confidence": "0..1 float",
                "reason": "short reason",
            },
            "constraints": [
                "Return JSON only, no markdown",
                "If unsure, set is_application_form=false",
            ],
        }

        try:
            response = self._client.responses.create(
                model=settings.doc_classifier_model,
                input=[
                    {
                        "role": "system",
                        "content": (
                            "You classify tender documents. "
                            "Be conservative: only mark as application form if clearly intended for supplier input."
                        ),
                    },
                    {
                        "role": "user",
                        "content": json.dumps(payload, ensure_ascii=False),
                    },
                ],
                max_output_tokens=max(80, settings.doc_classifier_max_output_tokens),
            )
            parsed = _parse_json(response.output_text.strip())
        except Exception as exc:
            return DocClassification(False, 0.0, f"LLM classification failed: {exc}")

        return DocClassification(
            is_application_form=_coerce_bool(parsed.get("is_application_form")),
            confidence=_coerce_confidence(parsed.get("confidence"), fallback=0.0),
            reason=str(parsed.get("reason") or "LLM classification").strip(),
        )
