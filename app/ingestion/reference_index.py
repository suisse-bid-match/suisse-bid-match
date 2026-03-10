from __future__ import annotations

import html
import re
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.models import DocumentInfo, ReferenceChunk


def _clean_text(text: str) -> str:
    text = html.unescape(text)
    text = text.replace("\u2002", " ").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _chunk_paragraphs(paragraphs: list[str], max_chars: int = 900) -> list[str]:
    chunks: list[str] = []
    buffer: list[str] = []
    size = 0
    for paragraph in paragraphs:
        p = paragraph.strip()
        if not p:
            continue
        next_size = size + len(p) + (1 if buffer else 0)
        if buffer and next_size > max_chars:
            chunks.append(" ".join(buffer).strip())
            buffer = [p]
            size = len(p)
            continue
        buffer.append(p)
        size = next_size
    if buffer:
        chunks.append(" ".join(buffer).strip())
    return [chunk for chunk in chunks if chunk]


def _docx_paragraphs(path: Path) -> list[str]:
    with ZipFile(path) as zf:
        xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    xml = re.sub(r"</w:(?:p|tr|tc)>", "\n", xml)
    xml = re.sub(r"<[^>]+>", "", xml)
    return [line for line in (_clean_text(x) for x in xml.splitlines()) if line]


def _xlsx_paragraphs(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values: list[str] = []
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    val = _clean_text(cell.value)
                else:
                    val = str(cell.value)
                if val:
                    values.append(val)
            if values:
                lines.append(f"[{ws.title}] " + " | ".join(values))
    return lines


def _pdf_chunks(path: Path, doc: DocumentInfo) -> list[ReferenceChunk]:
    reader = PdfReader(str(path))
    chunks: list[ReferenceChunk] = []
    idx = 0
    for page_no, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        lines = [_clean_text(x) for x in text.splitlines() if _clean_text(x)]
        for chunk_text in _chunk_paragraphs(lines):
            chunk_id = f"{doc.doc_id}:chunk:{idx:04d}"
            idx += 1
            chunks.append(
                ReferenceChunk(
                    chunk_id=chunk_id,
                    doc_id=doc.doc_id,
                    doc_name=doc.name,
                    section_path=f"page:{page_no}",
                    page_or_anchor=str(page_no),
                    text=chunk_text,
                    tokens=len(chunk_text.split()),
                )
            )
    return chunks


def build_reference_chunks(doc: DocumentInfo, file_path: Path) -> list[ReferenceChunk]:
    if doc.role != "REFERENCE_ONLY":
        return []

    chunks: list[ReferenceChunk] = []
    raw_chunks: list[str] = []
    if doc.kind in {"docx", "docm"}:
        raw_chunks = _chunk_paragraphs(_docx_paragraphs(file_path))
    elif doc.kind == "xlsx":
        raw_chunks = _chunk_paragraphs(_xlsx_paragraphs(file_path))
    elif doc.kind == "pdf":
        return _pdf_chunks(file_path, doc)

    for idx, chunk_text in enumerate(raw_chunks):
        chunk_id = f"{doc.doc_id}:chunk:{idx:04d}"
        chunks.append(
            ReferenceChunk(
                chunk_id=chunk_id,
                doc_id=doc.doc_id,
                doc_name=doc.name,
                section_path="body",
                page_or_anchor=f"chunk:{idx}",
                text=chunk_text,
                tokens=len(chunk_text.split()),
            )
        )
    return chunks
