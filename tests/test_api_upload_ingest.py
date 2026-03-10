from __future__ import annotations

import mimetypes
import os
import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


def _uster_source() -> Path:
    return Path(
        os.getenv(
            "USTER_SOURCE_PATH",
            "/home/daz/all_things_for_genai_hackathon/real_tenders/20260220_Upload_simap_BKP_233-20260306T200217Z-3-001/20260220_Upload_simap_BKP_233",
        )
    )


@pytest.mark.skipif(not _uster_source().exists(), reason="Uster source package not found")
def test_api_ingest_accepts_multipart_upload_files() -> None:
    client = TestClient(app)
    files_payload: list[tuple[str, tuple[str, object, str]]] = []
    handles = []

    for source_file in sorted(_uster_source().iterdir()):
        if not source_file.is_file():
            continue
        mime = mimetypes.guess_type(source_file.name)[0] or "application/octet-stream"
        fh = source_file.open("rb")
        handles.append(fh)
        files_payload.append(("files", (source_file.name, fh, mime)))

    try:
        response = client.post("/api/packages/ingest", files=files_payload)
    finally:
        for fh in handles:
            fh.close()

    assert response.status_code == 200
    payload = response.json()
    assert payload["document_count"] == 7
    fields_resp = client.get(f"/api/packages/{payload['package_id']}/fields")
    assert fields_resp.status_code == 200
    assert payload["field_count"] == 0


def test_api_ingest_accepts_folder_style_relative_paths(tmp_path: Path) -> None:
    book_path = tmp_path / "Angebot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Firma"
    ws["B1"] = None
    wb.save(book_path)

    client = TestClient(app)
    with book_path.open("rb") as fh:
        response = client.post(
            "/api/packages/ingest",
            files=[("files", ("bid_pack/forms/Angebot.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document_count"] == 1

    fields_resp = client.get(f"/api/packages/{payload['package_id']}/fields")
    assert fields_resp.status_code == 200
    docs = fields_resp.json()["documents"]
    assert len(docs) == 1
    assert docs[0]["relative_path"].endswith("source/bid_pack/forms/Angebot.xlsx")


def test_api_ingest_blocks_path_traversal_filename(tmp_path: Path) -> None:
    book_path = tmp_path / "Angebot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Firma"
    ws["B1"] = None
    wb.save(book_path)

    client = TestClient(app)
    with book_path.open("rb") as fh:
        response = client.post(
            "/api/packages/ingest",
            files=[("files", ("../Angebot.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )

    assert response.status_code == 400
    assert "invalid upload path" in str(response.json()["detail"]).lower()


def test_api_ingest_start_reports_processing_progress(tmp_path: Path) -> None:
    book_path = tmp_path / "Angebot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Firma"
    ws["B1"] = None
    wb.save(book_path)

    client = TestClient(app)
    with book_path.open("rb") as fh:
        start = client.post(
            "/api/packages/ingest/start",
            files=[("files", ("pack/forms/Angebot.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
    assert start.status_code == 200
    payload = start.json()
    job_id = payload["job_id"]
    assert payload["uploaded_files"] == 1

    final_job = None
    saw_total = False
    for _ in range(80):
        resp = client.get(f"/api/packages/ingest/{job_id}")
        assert resp.status_code == 200
        job = resp.json()
        if job.get("total_files", 0) > 0:
            saw_total = True
        if job["status"] in {"completed", "failed"}:
            final_job = job
            break
        time.sleep(0.1)

    assert final_job is not None
    assert final_job["status"] == "completed"
    assert saw_total
    assert final_job["processed_files"] >= 1
    assert final_job["total_files"] >= 1
    assert final_job["package_id"]
